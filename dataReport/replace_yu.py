import pandas as pd  
import xlrd
import openpyxl

def  replace_SZ_radio(yxn_xlsx,xone_xls,yuegao_xls): 
    # 读取Excel文件  
    df_excel_yg = pd.read_excel(yuegao_xls,usecols=[1,6,7,11],engine='xlrd')  
    df_excel_xone = pd.read_excel(xone_xls,usecols=[3,31],engine='xlrd')  
    #df_excel.columns = ['id', 'ratio']  # 设置列名

    replace_dict_yg = {} 
    replace_dict_xone = {} 
    print(df_excel_yg)
    # **********************************
    print(df_excel_xone)


    for index, row in df_excel_yg.iterrows():
        #循环生成字典
        print(row)
        id, ratio,zy_num,zy_balance=str(row[0]),row[1],row[2],row[3]        
        replace_dict_yg[id] = {'ratio_yg': ratio,'zy_num':zy_num,'zy_balance':zy_balance}

    for index, row in df_excel_xone.iterrows():
        #循环生成字典
        print(row)
        id, ratio=str(row[0]),row[1]      
        replace_dict_xone[id] = {'ratio_xone': ratio}
    

 
    ##print("ok")




    workbook = openpyxl.load_workbook(yxn_xlsx) 
    worksheet=workbook['自有资金合约'] 
    for row in worksheet.iter_rows(min_row=2):
        id=str(row[17].value)
        if id in  replace_dict_yg:
            print(id+ "in replace_dict_yg")
            replacements_yg = replace_dict_yg[id]
            row[14].value = replacements_yg['ratio_yg']
            row[5].value = round(replacements_yg['zy_num']/10000,4)
            row[9].value = round(replacements_yg['zy_balance']/10000,4)
        
        if  id in  replace_dict_xone:
            print(id + "in replace_dict_xone")
            replacements_xone = replace_dict_xone[id]
            row[13].value = replacements_xone['ratio_xone']
    
    worksheet=workbook['资管纾困计划（自有出资）合约'] 
    for row in worksheet.iter_rows(min_row=2):
        id=str(row[17].value)
        if id in  replace_dict_yg:
            print("id in replace_dict_yg")
            replacements_yg = replace_dict_yg[id]
            row[14].value = replacements_yg['ratio_yg']
            row[5].value = round(replacements_yg['zy_num']/10000,4)
            row[9].value = round(replacements_yg['zy_balance']/10000,4)
        if  id in  replace_dict_xone:
            print("id in replace_dict_xone")
            replacements_xone = replace_dict_xone[id]
            row[13].value = replacements_xone['ratio_xone']    
        print(row[2])
        print("ok")
    
    worksheet=workbook['违约项目（自有资金合约）'] 
    for row in worksheet.iter_rows(min_row=2):
        id=str(row[17].value)
        if id in  replace_dict_yg:
            print("id in replace_dict_yg")
            replacements_yg = replace_dict_yg[id]
            row[14].value = replacements_yg['ratio_yg']
            row[5].value = round(replacements_yg['zy_num']/10000,4)
            row[9].value = round(replacements_yg['zy_balance']/10000,4)
        if  id in  replace_dict_xone:
            print("id in replace_dict_xone")
            replacements_xone = replace_dict_xone[id]
            row[13].value = replacements_xone['ratio_xone']    

    workbook.save(yxn_xlsx)




replace_SZ_radio('Y:/报表_于晓妮/股票质押回购合约明细-yxn.xlsx','Y:/报表_于晓妮/xone.xls','Y:/报表_于晓妮/yg_20241129.xls')

